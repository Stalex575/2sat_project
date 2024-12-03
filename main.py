"""
2SAT project
"""
import argparse
import pandas as pd

def read_from_terminal() -> tuple[str, str, list[int]]:
    """
    Reads input data from the command line and returns it as a tuple.

    Returns:
    - tuple[str, str, str, list[int]]
    """
    parser = argparse.ArgumentParser()
    parser.add_argument('--f', type=str, required=True)
    parser.add_argument('--m', type=str, required=True)
    parser.add_argument('--r', type=str, required=True)
    parser.add_argument('--c', type=str, required=True)
    args = parser.parse_args()
    id_users = [int(x) for x in args.c.split(',')]
    return (args.f, args.m, args.r, id_users)

def read_mods(filename:str) -> dict:
    """
Reads a file containing modification data and returns a dictionary
where the key is the modification mod_id and the value is a tuple (name, user_visibility).
Return dictionary {key(mod_id): (name,user_visibility)}
:param filename: The path to the file containing the modification data.
:return modifications_dict: A dictionary where the key is mod_id and the value\
is a tuple (name: str, user_visibility: int).
    """
    with open(filename, encoding='utf-8') as f:
        modifications_dict = {}
        f.readline()
        for i in f:
            i = i.strip().split(';')
            if len(i) != 3:
                print('Неправильно написані дані, будь ласка, відкоригуйте файл!')
                break
            mod_id, name, user_visibility = i
            modifications_dict[int(mod_id)] = (name.strip(),int(user_visibility))
    return modifications_dict

def read_graph(filename: str) -> dict[int: list[int]]:
    """
    Reads a file containing constraints and returns a dictionary where the key is the mod_id
    and the value is a list of submods.
    
    Parameters:
    - filename (str): The path to the file containing the constraints.
    
    Returns:
    - dict[int: list[int]]: A dictionary where the key is the 
    mod_id and the value is a list of submods.
    """
    constraints = {}
    with open(filename, mode='r', encoding='utf-8') as f:
        f.readline()
        for line in f:
            line = line.strip().split(';')
            mod_id = int(line[0])
            conflicts = line[1].strip().split(',')
            requirements = line[2].strip().split(',')
            constraints[mod_id] = ([-int(num) for num in conflicts \
            if conflicts != ['']] + [int(num) for num in requirements if line[2] != ''])
    return constraints

def read_exel_mods(filename: str, sheet_name: str) -> dict:
    """
    Reads a specific sheet from an Excel file containing modification data 
    and returns a dictionary where the key is the modification mod_id 
    and the value is a tuple (name, user_visibility).
    
    :param filename: The path to the Excel file containing the modification data.
    :param sheet_name: The name of the sheet to read from the Excel file.
    :return: A dictionary {key(mod_id): (name: str, user_visibility: int)}.

    >>> import pandas as pd
    >>> data = {'id': [1, 2], 'mods': ['Mod A', 'Mod B'], 'user_visibility': [1, 0]}
    >>> df = pd.DataFrame(data)
    >>> df.to_excel('test_mods.xlsx', sheet_name='Modifications', index=False)
    >>> read_exel_mods('test_mods.xlsx', 'Modifications')
    {1: ('Mod A', 1), 2: ('Mod B', 0)}
    """
    data = pd.read_excel(filename, sheet_name=sheet_name)
    required_columns = {'id', 'mods', 'user_visibility'}
    if not required_columns.issubset(data.columns):
        raise ValueError(f"Сторінка '{sheet_name}' повинна містити колонки: {required_columns}")
    modifications_dict = {
        row['id']: (row['mods'], row['user_visibility'])
        for _, row in data.iterrows()
    }
    return modifications_dict

def read_constraints(filename: str, sheet_name: str) -> dict[int, list[int]]:
    """
    Reads a specific sheet from an Excel file containing constraint data 
    and returns a dictionary where the key is the mod_id and the value 
    is a list of submods (conflicts and requirements).
    """
    
    data = pd.read_excel(filename, sheet_name=sheet_name)
    required_columns = {'id', 'conflicting id', 'must id'}
    if not required_columns.issubset(data.columns):
        raise ValueError(f"Sheet '{sheet_name}' must contain columns: {required_columns}")
    constraints_dict = {}
    for _, row in data.iterrows():
        mod_id = int(row['id'])
        conflicts = row.get('conflicting id', '')
        if pd.isna(conflicts): 
            conflicts_list = []
        elif isinstance(conflicts, str):
            conflicts_list = [-int(x.strip()) for x in conflicts.split(',') if x.strip()]
        else:
            conflicts_list = [-int(conflicts)]
        requirements = row.get('must id', '')
        if pd.isna(requirements): 
            requirements_list = []
        elif isinstance(requirements, str):
            requirements_list = [int(x.strip()) for x in requirements.split(',') if x.strip()]
        else:
            requirements_list = [int(requirements)]
        constraints_dict[mod_id] = conflicts_list + requirements_list
    
    return constraints_dict

import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

def write_modifications_to_excel(filename: str, sheet_name: str, input_data: str):
    """
    Parses the input data, processes compatible, required, and incompatible modifications, and writes them to an Excel file.
    Adds color formatting to the header text for better readability.
    
    :param filename: Path to the Excel file.
    :param sheet_name: Name of the new sheet to add.
    :param input_data: Input string containing modifications in a specific format.
    """
    try:
        required_format = r"Модифікації:.*сумісні.*Необхідні модифікації та підмодифікації:.*"
        if not re.match(required_format, input_data):
            print(input_data)
            return
        compatible = []
        required = []
        compatible_match = re.search(r"Модифікації:\s*(.*?)\s*(необхідні|$)", input_data, re.DOTALL)
        if compatible_match:
            compatible_data = compatible_match.group(1).strip()
            compatible = [(int(mod.split(' - ')[0]), mod.split(' - ')[1].strip()) for mod in compatible_data.split(',') if mod]
        required_match = re.search(r"Необхідні модифікації та підмодифікації:\s*(.*)", input_data, re.DOTALL)
        if required_match:
            required_data = required_match.group(1).strip()
            required = [(int(mod.split(' - ')[0]), mod.split(' - ')[1].strip()) for mod in required_data.split(',') if mod]
        try:
            workbook = load_workbook(filename)
        except FileNotFoundError:
            workbook = Workbook()
            if "Sheet" in workbook.sheetnames:
                del workbook["Sheet"]

        if sheet_name in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' already exists in '{filename}'.")
        sheet = workbook.create_sheet(title=sheet_name)
        current_row = 1
        header_font = Font(bold=True, color="FF0000")  
        if compatible:
            sheet.cell(row=current_row, column=1, value="Сумісні модифікації:")
            sheet.cell(row=current_row, column=1).font = header_font
            current_row += 1
            for mod_id, description in compatible:
                sheet.cell(row=current_row, column=1, value=mod_id)
                sheet.cell(row=current_row, column=2, value=description)
                current_row += 1
        if required:
            sheet.cell(row=current_row, column=1, value="Необхідні модифікації та підмодифікації:")
            sheet.cell(row=current_row, column=1).font = header_font
            current_row += 1
            for mod_id, description in required:
                sheet.cell(row=current_row, column=1, value=mod_id)
                sheet.cell(row=current_row, column=2, value=description)
                current_row += 1
            workbook.save(filename)
        print(f"Data successfully written to '{sheet_name}' in '{filename}'.")

    except Exception as e:
        print(input_data)


def satisfy(graph: dict[int, list[int]], user_choice: list[int],\
    all_mods: dict[int: (str, int)]) -> dict[int, bool]:
    """
    Determines whether the user's choice of modifications is compatible with the constraints.
    
    Parameters:
    - graph (dict[int, list[int]]): A dictionary where the key is the mod_id and
    the value is a list of submods.
    - user_choice (list[int]): A list of mod_ids that the user has selected.
    - all_mods (dict[int: (str, int)]): A dictionary where the key is the mod_id
    and the value is a tuple (name, user_visibility).
    
    Returns:
    - dict[int, bool]: A dictionary where the key is the mod_id and the value is a boolean
    indicating whether the mod is compatible with the constraints.
    """

    def handle_submods(submod: int, use_modifications: dict[int, bool], graph: dict) -> None:
        """
        Recursively handles submods of a mod_id.
        
        Parameters:
        - submod (int): The mod_id of the submod.
        - use_modifications (dict[int, bool]): A dictionary where the key is the mod_id
        and the value is a boolean indicating whether the mod is compatible with the constraints.
        - graph (dict): A dictionary where the key is the mod_id and the value is a list of submods.
        """
        for mod_id in graph[submod]:
            # if the submod is required and is unset in the dict or is set to True, set it to True
            if mod_id > 0 and (use_modifications[abs(mod_id)] is None or\
                use_modifications[abs(mod_id)] is True):
                use_modifications[abs(mod_id)] = True
                handle_submods(abs(mod_id), use_modifications, graph)
            # if the submod is conflicting or if it is already set to False
            elif mod_id < 0 and (use_modifications[abs(mod_id)] is False or \
            use_modifications[abs(mod_id)] is None):
                use_modifications[abs(mod_id)] = False

            else:
                raise ValueError('Incompatible modifications.')

    use_modifications = {mod_id: None for mod_id in graph.keys()}
    for mod in user_choice:
        use_modifications[mod] = True
        handle_submods(mod, use_modifications, graph)
    for mod_id, properties in all_mods.items():
        # if value is unset and the mod is visible to the user
        if use_modifications[mod_id] is None and properties[1] == 1:
            use_modifications[mod_id] = False
    return use_modifications

def main():
    """
    Main function that reads the input data from the terminal, reads the modifications
    and constraints from the files, and determines whether the user's choice of modifications
    is compatible with the constraints.
    
    Returns:
    - str: A string indicating whether the user's choice of modifications is compatible
    with the constraints.
    """
    combined_file, modifications_file, restrictions_file, user_input = read_from_terminal()
    mods_dict = read_exel_mods(combined_file, modifications_file)
    try:
        mods = satisfy(read_constraints(combined_file, restrictions_file), user_input, mods_dict)
        mods = list(filter(lambda x: mods[x] is True, mods))
        mods_to_return = [f"{i} - {mods_dict[i][0]}" for i in mods]
        return f'Модифікації: {', '.join(el for el in [f"{i} - {mods_dict[i][0]}" for i in user_input])} сумісні. Необхідні \
модифікації та підмодифікації: {', '.join(el for el in mods_to_return)}'
    except ValueError:
        return f'Модифікації: {', '.join(el for el in [f"{i} - {mods_dict[i][0]}" for i in user_input])} несумісні'
write_modifications_to_excel("combined_modifications.xlsx", "result", main())
